from decimal import Decimal
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError
from django.db.models import Avg, Max, Min
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from apps.schools.models import ClassGroup
from .forms import ExamForm, MultiClassGroupExamApplicationForm
from .models import Exam, ExamApplication, StudentAnswer, StudentExamResult


def _get_selected_application(exam, application_id):
    if not application_id:
        return None

    return get_object_or_404(
        exam.applications.select_related("class_group"),
        pk=application_id,
    )


def _get_results_queryset(exam, selected_application=None):
    results = StudentExamResult.objects.filter(exam_application__exam=exam)

    if selected_application:
        results = results.filter(exam_application=selected_application)

    return results.select_related(
        "student",
        "exam_application",
        "exam_application__class_group",
    ).order_by("exam_application__class_group__name", "student__name")


def _build_difficulty_data(exam, selected_application=None):
    answer_filters = {"result__exam_application__exam": exam}
    total_results = StudentExamResult.objects.filter(exam_application__exam=exam)

    if selected_application:
        answer_filters["result__exam_application"] = selected_application
        total_results = total_results.filter(exam_application=selected_application)

    total_results = total_results.count()
    difficulty_data = []

    for question_number in range(1, exam.question_count + 1):
        correct_count = StudentAnswer.objects.filter(
            question_number=question_number,
            is_correct=True,
            **answer_filters,
        ).count()
        accuracy = (correct_count / total_results * 100) if total_results else 0
        difficulty_data.append(
            {
                "question_number": question_number,
                "correct_count": correct_count,
                "accuracy": round(accuracy, 2),
                "difficulty": round(100 - accuracy, 2),
            }
        )

    return difficulty_data


def _build_application_summaries(exam):
    summaries = []

    for application in exam.applications.select_related("class_group").all():
        results = StudentExamResult.objects.filter(exam_application=application)
        stats = results.aggregate(
            average_score=Avg("score"),
            highest_score=Max("score"),
            lowest_score=Min("score"),
        )
        summaries.append(
            {
                "application": application,
                "student_count": results.count(),
                "average_score": stats["average_score"],
                "highest_score": stats["highest_score"],
                "lowest_score": stats["lowest_score"],
            }
        )

    return summaries


@login_required
def dashboard_view(request):
    total_exams = Exam.objects.count()
    total_results = StudentExamResult.objects.count()
    total_class_groups = ClassGroup.objects.count()
    latest_exams = Exam.objects.order_by("-application_date", "-created_at")[:5]

    performance_summary = StudentExamResult.objects.aggregate(
        average_score=Avg("score"),
        highest_score=Max("score"),
        lowest_score=Min("score"),
    )

    context = {
        "total_exams": total_exams,
        "total_results": total_results,
        "total_class_groups": total_class_groups,
        "latest_exams": latest_exams,
        "performance_summary": performance_summary,
    }
    return render(request, "exams/dashboard.html", context)


@login_required
def exam_list_view(request):
    exams = Exam.objects.select_related("created_by").all().order_by("-application_date", "title")
    return render(request, "exams/exam_list.html", {"exams": exams})


@login_required
def exam_create_view(request):
    if request.method == "POST":
        form = ExamForm(request.POST)
        if form.is_valid():
            exam = form.save(commit=False)
            exam.created_by = request.user
            exam.save()
            messages.success(request, "Prova cadastrada com sucesso. Agora associe as turmas.")
            return redirect("exams:exam_application_manage", pk=exam.pk)
    else:
        form = ExamForm()

    return render(request, "exams/exam_form.html", {"form": form, "title": "Nova prova"})


@login_required
def exam_application_manage_view(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    applications = exam.applications.select_related("class_group").all()

    if request.method == "POST":
        form = MultiClassGroupExamApplicationForm(request.POST)
        if form.is_valid():
            class_groups = form.cleaned_data["class_groups"]
            notes = form.cleaned_data["notes"]
            is_active = form.cleaned_data["is_active"]

            created_count = 0
            skipped_count = 0

            for class_group in class_groups:
                try:
                    ExamApplication.objects.create(
                        exam=exam,
                        class_group=class_group,
                        notes=notes,
                        is_active=is_active,
                    )
                    created_count += 1
                except IntegrityError:
                    skipped_count += 1

            messages.success(
                request,
                f"Associacao concluida. {created_count} turma(s) vinculada(s) e {skipped_count} ja existia(m).",
            )
            return redirect("exams:exam_application_manage", pk=exam.pk)
    else:
        form = MultiClassGroupExamApplicationForm()

    context = {
        "exam": exam,
        "applications": applications,
        "form": form,
        "title": f"Aplicacoes - {exam.title}",
    }
    return render(request, "exams/exam_application_manage.html", context)


@login_required
def correction_select_student_view(request, application_id):
    exam_application = get_object_or_404(
        ExamApplication.objects.select_related("exam", "class_group"),
        pk=application_id,
    )
    students = exam_application.class_group.students.filter(is_active=True).order_by("name")
    corrected_student_ids = set(
        StudentExamResult.objects.filter(exam_application=exam_application).values_list("student_id", flat=True)
    )

    return render(
        request,
        "exams/correction_select_student.html",
        {
            "exam_application": exam_application,
            "exam": exam_application.exam,
            "students": students,
            "corrected_student_ids": corrected_student_ids,
        },
    )


@login_required
def correction_student_view(request, application_id, student_id):
    exam_application = get_object_or_404(
        ExamApplication.objects.select_related("exam", "class_group"),
        pk=application_id,
    )
    exam = exam_application.exam
    student = get_object_or_404(exam_application.class_group.students, pk=student_id)

    result, _ = StudentExamResult.objects.get_or_create(
        exam_application=exam_application,
        student=student,
        defaults={
            "corrected_by": request.user,
            "corrected_at": timezone.now(),
        },
    )

    existing_answers = {answer.question_number: answer for answer in result.answers.all()}
    question_numbers = list(range(1, exam.question_count + 1))

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "mark_all_correct":
            for question_number in question_numbers:
                answer = existing_answers.get(question_number)
                if answer:
                    answer.is_correct = True
                    answer.save(update_fields=["is_correct", "updated_at"])
                else:
                    StudentAnswer.objects.create(
                        result=result,
                        question_number=question_number,
                        is_correct=True,
                    )

        elif action == "clear_all":
            for question_number in question_numbers:
                answer = existing_answers.get(question_number)
                if answer:
                    answer.is_correct = False
                    answer.save(update_fields=["is_correct", "updated_at"])
                else:
                    StudentAnswer.objects.create(
                        result=result,
                        question_number=question_number,
                        is_correct=False,
                    )

        else:
            for question_number in question_numbers:
                is_correct = request.POST.get(f"question_{question_number}") == "1"

                answer = existing_answers.get(question_number)
                if answer:
                    answer.is_correct = is_correct
                    answer.save(update_fields=["is_correct", "updated_at"])
                else:
                    StudentAnswer.objects.create(
                        result=result,
                        question_number=question_number,
                        is_correct=is_correct,
                    )

        total_correct = result.answers.filter(is_correct=True).count()
        percentage = (
            (Decimal(total_correct) / Decimal(exam.question_count)) * Decimal("100")
            if exam.question_count
            else Decimal("0")
        )
        score = (
            (Decimal(total_correct) / Decimal(exam.question_count)) * exam.max_score
            if exam.question_count
            else Decimal("0")
        )

        result.total_correct = total_correct
        result.percentage = percentage.quantize(Decimal("0.01"))
        result.score = score.quantize(Decimal("0.01"))
        result.corrected_by = request.user
        result.corrected_at = timezone.now()
        result.save()

        messages.success(request, f"Correcao de {student.name} salva com sucesso.")
        if action == "save_and_back" or not action:
            return redirect("exams:correction_select_student", application_id=exam_application.pk)
        return redirect("exams:correction_student", application_id=exam_application.pk, student_id=student.pk)

    answers_map = {answer.question_number: answer.is_correct for answer in result.answers.all()}

    context = {
        "exam_application": exam_application,
        "exam": exam,
        "student": student,
        "result": result,
        "question_numbers": question_numbers,
        "answers_map": answers_map,
    }
    return render(request, "exams/correction_student.html", context)


@login_required
def exam_report_view(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    selected_application = _get_selected_application(exam, request.GET.get("application_id"))
    applications = exam.applications.select_related("class_group").all()
    results = _get_results_queryset(exam, selected_application)
    difficulty_data = _build_difficulty_data(exam, selected_application)
    ordered_difficulty = sorted(difficulty_data, key=lambda item: item["accuracy"])
    result_summary = results.aggregate(
        average_score=Avg("score"),
        highest_score=Max("score"),
        lowest_score=Min("score"),
    )

    context = {
        "exam": exam,
        "applications": applications,
        "selected_application": selected_application,
        "application_summaries": _build_application_summaries(exam),
        "results": results,
        "result_count": results.count(),
        "result_summary": result_summary,
        "difficulty_data": difficulty_data,
        "ordered_difficulty": ordered_difficulty,
    }
    return render(request, "exams/exam_report.html", context)


@login_required
def export_exam_results_excel_view(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    selected_application = _get_selected_application(exam, request.GET.get("application_id"))
    results = _get_results_queryset(exam, selected_application)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resultados"

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    bold_font = Font(bold=True)

    headers = (
        ["Turma", "Aluno"]
        + [f"Q{i}" for i in range(1, exam.question_count + 1)]
        + ["Total de acertos", "Nota", "Percentual"]
    )
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = bold_font
        cell.fill = header_fill

    question_totals = {i: 0 for i in range(1, exam.question_count + 1)}

    for result in results:
        answers_map = {answer.question_number: 1 if answer.is_correct else 0 for answer in result.answers.all()}
        row = [result.exam_application.class_group.name, result.student.name]

        for question_number in range(1, exam.question_count + 1):
            value = answers_map.get(question_number, 0)
            row.append(value)
            question_totals[question_number] += value

        row.extend([result.total_correct, float(result.score), float(result.percentage)])
        sheet.append(row)

    footer_total = ["", "Total por questao"] + [question_totals[i] for i in range(1, exam.question_count + 1)] + ["", "", ""]
    footer_percent = ["", "% de acerto"]

    total_students = results.count() or 1
    for question_number in range(1, exam.question_count + 1):
        pct = round((question_totals[question_number] / total_students) * 100, 2)
        footer_percent.append(pct)

    footer_percent += ["", "", ""]

    sheet.append(footer_total)
    sheet.append(footer_percent)

    for row_index in [sheet.max_row - 1, sheet.max_row]:
        for cell in sheet[row_index]:
            cell.font = bold_font
            cell.fill = header_fill

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            value = str(cell.value) if cell.value is not None else ""
            if len(value) > max_length:
                max_length = len(value)
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 30)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    file_suffix = f"_turma_{selected_application.class_group.pk}" if selected_application else ""
    response["Content-Disposition"] = f'attachment; filename="resultado_{exam.pk}{file_suffix}.xlsx"'
    return response


@login_required
def export_exam_summary_excel_view(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    selected_application = _get_selected_application(exam, request.GET.get("application_id"))
    results = StudentExamResult.objects.filter(exam_application__exam=exam)

    if selected_application:
        results = results.filter(exam_application=selected_application)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resumo por questao"

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    bold_font = Font(bold=True)

    headers = ["Questao", "Total de acertos", "% de acerto", "% de dificuldade"]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = bold_font
        cell.fill = header_fill

    total_students = results.count() or 1

    for question_number in range(1, exam.question_count + 1):
        answer_filters = {
            "result__exam_application__exam": exam,
            "question_number": question_number,
            "is_correct": True,
        }
        if selected_application:
            answer_filters["result__exam_application"] = selected_application

        correct_count = StudentAnswer.objects.filter(**answer_filters).count()
        accuracy = round((correct_count / total_students) * 100, 2)
        difficulty = round(100 - accuracy, 2)

        sheet.append([
            question_number,
            correct_count,
            accuracy,
            difficulty,
        ])

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            value = str(cell.value) if cell.value is not None else ""
            if len(value) > max_length:
                max_length = len(value)
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 28)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    file_suffix = f"_turma_{selected_application.class_group.pk}" if selected_application else ""
    response["Content-Disposition"] = f'attachment; filename="resumo_{exam.pk}{file_suffix}.xlsx"'
    return response


@login_required
def exam_update_view(request, pk):
    exam = get_object_or_404(Exam, pk=pk)

    if request.method == "POST":
        form = ExamForm(request.POST, instance=exam)
        if form.is_valid():
            form.save()
            messages.success(request, "Prova atualizada com sucesso.")
            return redirect("exams:exam_list")
    else:
        form = ExamForm(instance=exam)

    return render(request, "exams/exam_form.html", {"form": form, "title": f"Editar prova: {exam.title}"})
